
import json
import openpyxl
import requests
from openpyxl.styles import Alignment
from .odbeimport import *


def kvexporttoxls(path):
    url = 'http://cleaneffect.hu/ODBE_odbe_rk_betolt.php'
    #url = 'http://cleaneffect.hu/test.php'
    with urllib.request.urlopen(url) as url:
        data = json.loads(url.read().decode())

    kvLastColumn = 6
    wb = openpyxl.load_workbook(path)
    for name in wb.sheetnames:  # Az excel lapok száma szerint
        for column_data in wb[name]['A']:
            ndnToFind = column_data.value
            filtered_list = [
                dictionary for dictionary in data
                if dictionary['ndn'] == ndnToFind
            ]

            if filtered_list != []:
                if filtered_list[0]['kv0'] != '':
                    wb[name].cell(row=column_data.row, column=(kvLastColumn)).value = int(filtered_list[0]['kv0'])
                    wb[name].cell(row=column_data.row, column=kvLastColumn).alignment = Alignment(horizontal="center")
                if filtered_list[0]['kv1'] != '':
                    wb[name].cell(row=column_data.row, column=(kvLastColumn - 1)).value = int(filtered_list[0]['kv1'])
                    wb[name].cell(row=column_data.row, column=kvLastColumn - 1).alignment = Alignment(horizontal="center")
                if filtered_list[0]['kv2'] != '':
                    wb[name].cell(row=column_data.row, column=(kvLastColumn - 2)).value = int(filtered_list[0]['kv2'])
                    wb[name].cell(row=column_data.row, column=kvLastColumn - 2).alignment = Alignment(horizontal="center")
                if filtered_list[0]['kv3'] != '':
                    wb[name].cell(row=column_data.row, column=(kvLastColumn - 3)).value = int(filtered_list[0]['kv3'])
                    wb[name].cell(row=column_data.row, column=kvLastColumn- 3).alignment = Alignment(horizontal="center")

    wb.save('/Users/futurex/Desktop/' + str(path))

# ----- RAKTÁRKÉSZLETET A NYOMKÖVETŐBE EMELI -----
def rkexportotxls(path):
    url = 'http://cleaneffect.hu/ODBE_rk_excelbe.php'
    with urllib.request.urlopen(url) as url:
        data = json.loads(url.read().decode())

    wb = openpyxl.load_workbook(path)
    for name in wb.sheetnames:  # Az excel lapok száma szerint
        for column_data in wb[name]['A']:
            ndnToFind = column_data.value
            filtered_list = [
                dictionary for dictionary in data
                if dictionary['ndn'] == ndnToFind and dictionary['ean'] != ""
            ]
            if filtered_list != []:
                if filtered_list[0]['mennyiseg'] != '':
                    wb[name].cell(row=column_data.row, column=8).value = int(filtered_list[0]['mennyiseg'])
                    wb[name].cell(row=column_data.row, column=8).alignment = Alignment(horizontal="center")


    wb.save('/Users/futurex/Desktop/' + str(path))

