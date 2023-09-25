import urllib
import openpyxl
from collections import OrderedDict
import json
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ===== E betűs NDN azonosítókat a nyomkövető excelbe exportálja =====
def odbeeexport(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    last_row = sheet.max_row
    # Betölti az ODBE szótárt
    url = 'http://cleaneffect.hu/ODBE_betolt.php'
    with urllib.request.urlopen(url) as url:
        data = json.loads(url.read().decode())

    i = 1
    while True:
        ertek = sheet.cell(row=1, column=i).value
        if ertek == "NDN_CODE" and i < 30:
            break
        i = i + 1

    for column_data in sheet['C']:
        filtered_list = [
            dictionary for dictionary in data
            if dictionary['trafikcikkod'] == str(column_data.value)
        ]

        if filtered_list != []:
            sheet.cell(row=column_data.row, column=i).value = filtered_list[0]['ndn']
    wb.save('/Users/futurex/Desktop/' + str(path))

