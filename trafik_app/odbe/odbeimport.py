import urllib
from collections import OrderedDict
import json
import openpyxl
import requests
from datetime import datetime
from datetime import timedelta


def odbekvinfobetolt():
    url = 'http://cleaneffect.hu/ODBE_KVinfo.php'
    with urllib.request.urlopen(url) as url:
        data = json.loads(url.read().decode())
    return_data = []
    for adat in data:
        KVinfo = OrderedDict()
        KVinfo['id'] = adat['id']
        KVinfo['datumkezdet'] = adat['datumkezdet']
        date_str = str(adat['datumkezdet'])
        date_object = datetime.strptime(date_str, '%Y-%m-%d').date()
        datumvege = date_object + timedelta(days=6)
        KVinfo['datumvege'] = str(datumvege)
        KVinfo['megjegyzes'] = adat['megjegyzes']
        return_data.append(KVinfo)
    return return_data


def odbekvimport(path, datum, megjegyzes):

    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    last_row = sheet.max_row
    keszletvaltozas = []
    for x in range(2, last_row + 1):
        keszletvaltozas_termek = OrderedDict()
        keszletvaltozas_termek['cikkszam'] = sheet.cell(row=x, column=3).value
        keszletvaltozas_termek['megnevezes'] = sheet.cell(row=x, column=5).value
        keszletvaltozas_termek['mennyiseg'] = sheet.cell(row=x, column=6).value
        keszletvaltozas.append(keszletvaltozas_termek)
    j = json.dumps(keszletvaltozas)  # lista json file-ba
    url = 'http://cleaneffect.hu/ODBE_keszletvaltozasment.php'
    thisdict = dict(datumkezdet=datum, megjegyzes=megjegyzes, json=j)
    x = requests.post(url, data=thisdict)
    print(x.text)

# ====== Raktárkészlet excel importálása adatbázisba ======
def odberkimport(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    last_row = sheet.max_row
    ndnOszlop = None
    eanOszlop = None
    i = 1
    while True:
        ertek = sheet.cell(row=1, column=i).value
        if i == 30: # ----- Amig megtalálja az NDN_CODE-ot a 30-as oszlopig
            break
        if ertek == 'NDN_CODE':
            ndnOszlop = i
        if ertek == 'EAN':
            eanOszlop = i
        i = i + 1
    raktar = []
    for x in range(2, last_row + 1):
        raktar_termek = OrderedDict()
        raktar_termek['trafikcikkod'] = sheet.cell(row=x, column=3).value
        raktar_termek['megnevezes'] = sheet.cell(row=x, column=4).value
        raktar_termek['mennyiseg'] = sheet.cell(row=x, column=6).value
        raktar_termek['ndn'] = sheet.cell(row=x, column=ndnOszlop).value
        raktar_termek['ean'] = sheet.cell(row=x, column=eanOszlop).value
        raktar.append(raktar_termek)

    j = json.dumps(raktar)  # lista json file-ba

    url = 'http://cleaneffect.hu/ODBE_raktarMent.php'
    x = requests.post(url, data=j)
    print(x.text)

# ====== Üres ODBE excel importálása adatbázisba ======
def odbeureslistaimport(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    last_row = sheet.max_row
    odbeLista = []
    for x in range(2, last_row + 1):
        odbelista_termek = OrderedDict()
        odbelista_termek['ndn'] = sheet.cell(row=x, column=1).value
        odbelista_termek['megnevezes'] = sheet.cell(row=x, column=2).value
        odbelista_termek['gyarto'] = sheet.cell(row=x, column=4).value
        odbelista_termek['kategoria'] = sheet.cell(row=x, column=5).value
        odbeLista.append(odbelista_termek)

    j = json.dumps(odbeLista)  # lista json file-ba

    url = 'http://cleaneffect.hu/ODBE_ureslistament.php'
    x = requests.post(url, data=j)
    print(x.text)
