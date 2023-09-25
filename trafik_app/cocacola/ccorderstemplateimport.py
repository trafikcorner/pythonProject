import itertools
import json
import urllib
import openpyxl
from openpyxl.styles import Alignment
from collections import OrderedDict
import requests

def ccotimport():
    print('CC orders template import')
    wb = openpyxl.load_workbook('/Users/futurex/Desktop/orders_template.xlsx')
    sheet = wb.active
    last_row = sheet.max_row
    keszletvaltozas = []
    for x in range(6, last_row + 1):
        keszletvaltozas_termek = OrderedDict()
        keszletvaltozas_termek['cikkszam'] = sheet.cell(row=x, column=1).value
        keszletvaltozas_termek['termeknev'] = sheet.cell(row=x, column=2).value
        keszletvaltozas_termek['marka'] = sheet.cell(row=x, column=3).value
        keszletvaltozas_termek['kiszereles'] = sheet.cell(row=x, column=4).value
        keszletvaltozas_termek['csomagolas'] = sheet.cell(row=x, column=5).value
        keszletvaltozas_termek['mennyiseg'] = sheet.cell(row=x, column=6).value
        keszletvaltozas_termek['mertekegyseg'] = sheet.cell(row=x, column=7).value
        keszletvaltozas_termek['trafikcikkszam'] = sheet.cell(row=x, column=8).value
        keszletvaltozas_termek['koteg'] = sheet.cell(row=x, column=9).value
        keszletvaltozas.append(keszletvaltozas_termek)
    j = json.dumps(keszletvaltozas)  # lista json file-ba
    url = 'http://cleaneffect.hu/COCA_ment.php'
    x = requests.post(url, data=j)
    print(x.text)

