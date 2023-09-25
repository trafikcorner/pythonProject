import itertools
import json
import urllib
import openpyxl
from openpyxl.styles import Alignment
from collections import OrderedDict
import requests

def ccrkexport_otbe():
    url = 'http://cleaneffect.hu/COCA_betolt.php'
    with urllib.request.urlopen(url) as url:
        data = json.loads(url.read().decode())
    wb = openpyxl.load_workbook('/Users/futurex/Desktop/orders_template.xlsx')
    sheet = wb.active
    sheet.cell(row=5, column=8).value = "Boltban db"
    sheet.cell(row=5, column=9).value = "Boltban köteg"
    sheet.cell(row=5, column=10).value = "Köteg"
    sheet.cell(row=5, column=11).value = "Trafikcikkszám"
    for cella in sheet['A']:
        cikkszamKeresendo = cella.value
        filtered_list = [
            dictionary for dictionary in data
            if dictionary['cikkszam'] == str(cikkszamKeresendo)
        ]
        if filtered_list != []:
            for adat in filtered_list:
                sheet.cell(row=cella.row, column=8).value = int(adat['raktarmennyiseg'])
                sheet.cell(row=cella.row, column=8).alignment = Alignment(horizontal="center")
                if int(adat['raktarmennyiseg']) != 0 and int(adat['koteg']) != 0:
                    sheet.cell(row=cella.row, column=9).value = int(adat['raktarmennyiseg']) / int(adat['koteg'])
                    sheet.cell(row=cella.row, column=9).alignment = Alignment(horizontal="center")
                sheet.cell(row=cella.row, column=10).value = int(adat['koteg'])
                sheet.cell(row=cella.row, column=10).alignment = Alignment(horizontal="center")
                sheet.cell(row=cella.row, column=11).value = int(adat['trafikcikkszam'])
                sheet.cell(row=cella.row, column=11).alignment = Alignment(horizontal="center")
    wb.save('/Users/futurex/Desktop/orders_template.xlsx')