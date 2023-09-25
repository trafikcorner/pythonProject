import itertools
import json
import urllib
import openpyxl
from openpyxl.styles import Alignment


### ----------------------------------------------------------
### Coco Cola orders template-be beemeli a készletváltozásokat
### ----------------------------------------------------------
def cckvexport():
    print("cc kv export")
    url = 'http://cleaneffect.hu/COCA_KV_excelbe.php'
    with urllib.request.urlopen(url) as url:
        data = json.loads(url.read().decode())
    # --- Dátumok tömbbe ---
    datumok = []
    for adat in data:
        if adat['datumkezdet'] not in datumok:
            datumok.append(adat['datumkezdet'])

    wb = openpyxl.load_workbook('/Users/futurex/Desktop/orders_template.xlsx')
    sheet = wb.active
    for cella in sheet['K']:
        cikkszamKeresendo = cella.value
        filtered_list = [
            dictionary for dictionary in data
            if dictionary['cikkszam'] == str(cikkszamKeresendo)
        ]
        if filtered_list != []:
            for adat in filtered_list:
                index = datumok.index(adat['datumkezdet'])
                sheet.cell(row=cella.row, column=12 + index).value = int(adat['mennyiseg'])
                sheet.cell(row=cella.row, column=12 + index).alignment = Alignment(horizontal="center")
    wb.save('/Users/futurex/Desktop/orders_template.xlsx')